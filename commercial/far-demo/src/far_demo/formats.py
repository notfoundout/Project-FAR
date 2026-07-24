from __future__ import annotations

import csv
import io
import json
import re
import tomllib
from pathlib import Path
from typing import Any

import xmltodict
import yaml
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

SUPPORTED_EXTENSIONS = {
    ".json",
    ".jsonl",
    ".yaml",
    ".yml",
    ".toml",
    ".xml",
    ".csv",
    ".txt",
    ".md",
    ".docx",
    ".pdf",
    ".xlsx",
}

ACCEPT_ATTRIBUTE = ",".join(sorted(SUPPORTED_EXTENSIONS))


def parse_package_file(filename: str, data: bytes) -> dict[str, Any]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported file type {suffix or '(none)'}. Supported: {supported}.")

    if suffix == ".json":
        return _require_mapping(json.loads(data.decode("utf-8")))
    if suffix == ".jsonl":
        objects = [json.loads(line) for line in data.decode("utf-8").splitlines() if line.strip()]
        if len(objects) != 1:
            raise ValueError("A FAR JSONL upload must contain exactly one non-empty JSON object.")
        return _require_mapping(objects[0])
    if suffix in {".yaml", ".yml"}:
        return _require_mapping(yaml.safe_load(data.decode("utf-8")))
    if suffix == ".toml":
        return _require_mapping(tomllib.loads(data.decode("utf-8")))
    if suffix == ".xml":
        return _normalize_xml(xmltodict.parse(data))
    if suffix == ".csv":
        return _table_rows_to_payload(csv.DictReader(io.StringIO(data.decode("utf-8"))))
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("The XLSX file is empty.")
        headers = [str(value or "").strip() for value in rows[0]]
        records = [dict(zip(headers, row, strict=False)) for row in rows[1:]]
        return _table_rows_to_payload(records)
    if suffix == ".docx":
        document = Document(io.BytesIO(data))
        return _parse_embedded_text("\n".join(paragraph.text for paragraph in document.paragraphs))
    if suffix == ".pdf":
        reader = PdfReader(io.BytesIO(data))
        return _parse_embedded_text("\n".join(page.extract_text() or "" for page in reader.pages))
    return _parse_embedded_text(data.decode("utf-8"))


def _require_mapping(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError("The uploaded file must decode to one FAR decision-package object.")
    return value


def _parse_embedded_text(text: str) -> dict[str, Any]:
    candidates = [text.strip()]
    candidates.extend(
        match.group(1).strip()
        for match in re.finditer(r"```(?:json|jsonl|yaml|yml|toml)?\s*(.*?)```", text, re.I | re.S)
    )
    errors: list[str] = []
    for candidate in candidates:
        if not candidate:
            continue
        for parser in (json.loads, yaml.safe_load, tomllib.loads):
            try:
                value = parser(candidate)
                if isinstance(value, dict):
                    return value
            except Exception as exc:  # parser-specific errors are summarized below
                errors.append(str(exc))
    raise ValueError(
        "No embedded FAR package was found. Text, Markdown, PDF, and DOCX files must contain "
        "one JSON, YAML, or TOML decision-package object."
    )


def _normalize_xml(data: dict[str, Any]) -> dict[str, Any]:
    root = next(iter(data.values())) if len(data) == 1 else data
    root = _require_mapping(root)
    payload = dict(root)

    nodes_container = payload.get("nodes", {})
    dependencies_container = payload.get("dependencies", {})
    nodes = nodes_container.get("node", []) if isinstance(nodes_container, dict) else nodes_container
    dependencies = (
        dependencies_container.get("dependency", [])
        if isinstance(dependencies_container, dict)
        else dependencies_container
    )
    payload["nodes"] = _as_list(nodes)
    payload["dependencies"] = _as_list(dependencies)

    requirements = payload.get("authorization_requirements", {})
    if isinstance(requirements, dict):
        requirements = requirements.get("node_id", requirements.get("requirement", []))
    payload["authorization_requirements"] = _as_list(requirements)

    unknowns = payload.get("unknowns", {})
    if isinstance(unknowns, dict):
        unknowns = unknowns.get("unknown", [])
    payload["unknowns"] = _as_list(unknowns)

    payload["trace_completeness"] = float(payload.get("trace_completeness", 0))
    payload.setdefault("metadata", {})
    payload.setdefault("proposed_action", {})
    for node in payload["nodes"]:
        if isinstance(node, dict):
            node.setdefault("attributes", {})
    return payload


def _as_list(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    return value if isinstance(value, list) else [value]


def _table_rows_to_payload(rows: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "nodes": [],
        "dependencies": [],
        "authorization_requirements": [],
        "unknowns": [],
        "metadata": {},
        "proposed_action": {},
    }
    for raw in rows:
        row = {str(key).strip(): value for key, value in dict(raw).items() if key is not None}
        record_type = str(row.get("record_type") or "").strip().lower()
        if not record_type:
            continue
        if record_type == "meta":
            key = str(row.get("key") or "").strip()
            value = _coerce(row.get("value"))
            if key.startswith("metadata."):
                payload["metadata"][key.removeprefix("metadata.")] = value
            elif key.startswith("proposed_action."):
                payload["proposed_action"][key.removeprefix("proposed_action.")] = value
            else:
                payload[key] = value
        elif record_type == "node":
            payload["nodes"].append(
                {
                    "node_id": str(row.get("node_id") or "").strip(),
                    "kind": str(row.get("kind") or "").strip(),
                    "statement": str(row.get("statement") or "").strip(),
                    "attributes": _coerce(row.get("attributes") or "{}"),
                }
            )
        elif record_type == "dependency":
            payload["dependencies"].append(
                {
                    "source_id": str(row.get("source_id") or "").strip(),
                    "target_id": str(row.get("target_id") or "").strip(),
                    "relation": str(row.get("relation") or "").strip(),
                }
            )
        elif record_type == "authorization_requirement":
            payload["authorization_requirements"].append(
                str(row.get("node_id") or row.get("value") or "").strip()
            )
        elif record_type == "unknown":
            payload["unknowns"].append(str(row.get("statement") or row.get("value") or "").strip())
        else:
            raise ValueError(f"Unknown tabular record_type: {record_type}.")
    return payload


def _coerce(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return ""
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text
